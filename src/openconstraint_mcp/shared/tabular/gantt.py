"""The cell-grid Gantt sheet: resolve a ``GanttSpec``, then render a timeline.

Split like ``style``: ``resolve_gantt`` validates and computes without touching
a workbook, so ``core`` can refuse a bad spec before staging a file, and
``render_gantt`` only draws what was resolved.

Times are discrete non-negative integers — the native shape of CP-SAT and
MiniZinc scheduling output. A float, a numeric string, or a null is refused
rather than coerced, matching this package's refuse-don't-coerce posture in
``guards``.

The grid has two independent axes and they name different columns: ``row_column``
decides which tasks share a row (the resource view), while ``color_column``
decides only what a bar is coloured by. A job shop wants both — rows per machine,
colour per job — so neither can stand in for the other. Both are keyed by the
label text they render, so two distinct values that would render alike are
refused rather than merged (see ``_check_distinct_label``).

Bar colours are the ``dataviz`` skill's reference categorical palette, in its
documented slot order. That order is the colour-blindness safety mechanism, so
it is never reordered. The CVD ΔE check run during design compared arbitrary
pairs across the first three slots only; beyond three colours the separation
rests on the palette's own published guarantees, not on anything measured here.
Colours cycle through the eight slots because the value set comes from the
caller's data and cannot be folded into an "Other" bucket the way a chart series
can; the legend prints every value's name beside its swatch, so identity never
rests on colour alone.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from ...schemas.tabular import GanttSpec, TabularCell
from .columns import column_index
from .guards import (
    _check_no_carriage_return,
    _check_no_illegal_xml_characters,
    _check_not_oversized_xlsx_string,
)
from .limits import GANTT_MAX_HORIZON_COLUMNS

# The dataviz reference palette's categorical slots, light mode, in order.
_COLOR_FILL_RGB: tuple[str, ...] = (
    "FF2A78D6",
    "FFEB6834",
    "FF1BAF7A",
    "FFEDA100",
    "FFE87BA4",
    "FF008300",
    "FF4A3AA7",
    "FFE34948",
)

# Narrow enough that a long horizon still reads as a timeline.
_TIME_COLUMN_WIDTH: int = 3

_TASK_COLUMN_HEADER: str = "Task"

# A null cell renders as an explicit placeholder rather than "": an empty string
# writes a degenerate cell openpyxl reads back as null (the same reason
# ``guards._reject_xlsx_empty_strings`` refuses one), and a blank-named legend
# swatch would put identity back on colour alone. Each names the missing DATA,
# not the channel it feeds — "(no colour)" beside a coloured swatch would read
# as a contradiction.
_MISSING_TASK_LABEL: str = "(untitled task)"
_MISSING_COLOR_NAME: str = "(uncategorized)"
_MISSING_ROW_LABEL: str = "(no group)"


@dataclass(frozen=True)
class ResolvedTask:
    """One task's bar: its half-open span, its fill colour, and the row it sits on.

    ``bar_label`` is the text drawn inside the bar, or ``None`` when column A
    already names this task — an ungrouped grid gives every task its own row, so
    a bar label there would only repeat what is one cell to its left.
    """

    bar_label: str | None
    start: int
    end: int
    color: str | None
    grid_row: int


@dataclass(frozen=True)
class ResolvedGantt:
    """Everything ``render_gantt`` needs, with nothing left to validate.

    ``row_labels`` is column A top to bottom, one entry per grid row, and
    ``row_header`` is the cell above it. Ungrouped they are the task labels
    under ``"Task"``; grouped they are the resource values under that column's
    own header, repeated once per sub-row a resource needed.
    """

    sheet_name: str
    title: str | None
    tasks: tuple[ResolvedTask, ...]
    row_header: str
    row_labels: tuple[str, ...]
    colors: tuple[str | None, ...]
    horizon: int


@dataclass(frozen=True)
class _Staged:
    """One task read off a row, before its grid row is known."""

    label: str
    start: int
    end: int
    color: str | None
    group: str | None


def _require_time(value: object, *, row_index: int, role: str) -> int:
    """Return ``value`` as a discrete time, or raise ``ValueError``."""
    if isinstance(value, bool) or not isinstance(value, int):
        raise ValueError(
            f"the {role} at row {row_index} is {value!r}; a Gantt needs a discrete "
            f"integer time, and values are never coerced"
        )
    return value


def _rendered_text(value: TabularCell, *, placeholder: str) -> str:
    """Render one cell as the label text a Gantt sheet shows."""
    return placeholder if value is None else str(value)


def _check_distinct_label(
    value: TabularCell,
    claimed: dict[str, TabularCell],
    *,
    placeholder: str,
    row_index: int,
    role: str,
) -> None:
    """Reject a value rendering as a label another value already claimed.

    Both axes key the grid by label text, so ``1`` and ``"1"`` — distinct cells
    the ``TabularCell`` union keeps apart precisely so a write never retypes
    them — would otherwise share one resource row and one colour category. Nor
    can the collision be rendered its way out of: a repeated row label already
    *means* something here (a spilled sub-row repeats its group's name), so two
    machines both named ``1`` read as one machine running overlapping work.
    A null is checked against the same map, since it renders as ``placeholder``
    and a cell holding that literal string claims the identical label.

    Refused rather than coerced or silently merged, as everywhere else in this
    package: ``claimed`` maps each label to the one value allowed to produce it.
    """
    label: str = _rendered_text(value, placeholder=placeholder)
    previous: TabularCell = claimed.setdefault(label, value)
    if previous != value:
        raise ValueError(
            f"the {role} at row {row_index} is {value!r}, but {previous!r} already renders "
            f"as the same label {label!r}; a Gantt names its rows and legend entries by that "
            f"text, so the two would be indistinguishable in the sheet — give them labels "
            f"that differ"
        )


def _assign_grid_rows(staged: list[_Staged]) -> tuple[list[ResolvedTask], list[str]]:
    """Pack tasks onto shared grid rows, one block of rows per group.

    Within a group each task takes the first sub-row already free at its start,
    opening a new one only when none is — the greedy interval-partitioning rule,
    which uses exactly as many sub-rows as the group's deepest overlap. A
    disjunctive resource (no two tasks at once, the job-shop case) therefore
    collapses to a single row, while a cumulative one spills only as far as it
    must. Overlap is never resolved by drawing one bar over another: this
    package refuses to lose data silently, and a dropped bar is lost data.

    A spilled sub-row repeats its group's name rather than being left blank, so
    a resource is never identified by position alone.
    """
    groups: dict[str | None, list[_Staged]] = {}
    for task in staged:
        groups.setdefault(task.group, []).append(task)

    tasks: list[ResolvedTask] = []
    row_labels: list[str] = []
    for group, members in groups.items():
        base: int = len(row_labels)
        sub_row_ends: list[int] = []
        for task in sorted(members, key=lambda member: member.start):
            index: int = next(
                (slot for slot, end in enumerate(sub_row_ends) if end <= task.start),
                len(sub_row_ends),
            )
            if index == len(sub_row_ends):
                sub_row_ends.append(task.end)
            else:
                sub_row_ends[index] = task.end
            tasks.append(
                ResolvedTask(
                    bar_label=task.label,
                    start=task.start,
                    end=task.end,
                    color=task.color,
                    grid_row=base + index,
                )
            )
        label: str = _rendered_text(group, placeholder=_MISSING_ROW_LABEL)
        row_labels.extend([label] * len(sub_row_ends))
    return tasks, row_labels


def resolve_gantt(
    headers: list[str], rows: list[list[TabularCell]], spec: GanttSpec
) -> ResolvedGantt:
    """Resolve ``spec`` against the table, or raise ``ValueError``.

    ``spec.title`` is free text, not a column reference, so it is not checked
    against ``headers``. It and ``spec.sheet_name`` are still user-controlled
    strings bound for the same XML writer as a cell, so both go through the
    same character guards the data sheet's own values do — the schema's
    sheet-name rules cover Excel's structural restrictions, not XML's. The
    title additionally lands in a real cell (``A1``), so it also gets the
    per-cell length and empty-string guards a data cell gets; a sheet name is
    not a cell and the schema already bounds it at 31 characters.
    """
    _check_no_illegal_xml_characters(spec.sheet_name, f"gantt sheet_name {spec.sheet_name!r}")
    _check_no_carriage_return(spec.sheet_name, f"gantt sheet_name {spec.sheet_name!r}")
    if spec.title is not None:
        if spec.title == "":
            raise ValueError(
                "the gantt title is an empty string, which renders as a blank first row "
                "while still shifting the grid down; omit title entirely instead of "
                "sending an empty string"
            )
        _check_no_illegal_xml_characters(spec.title, "the gantt title")
        _check_no_carriage_return(spec.title, "the gantt title")
        _check_not_oversized_xlsx_string(spec.title, "the gantt title")

    task_index = column_index(headers, spec.task_column, "task_column")
    start_index = column_index(headers, spec.start_column, "start_column")
    end_index = (
        None if spec.end_column is None else column_index(headers, spec.end_column, "end_column")
    )
    duration_index = (
        None
        if spec.duration_column is None
        else column_index(headers, spec.duration_column, "duration_column")
    )
    color_index = (
        None
        if spec.color_column is None
        else column_index(headers, spec.color_column, "color_column")
    )
    row_group_index = (
        None if spec.row_column is None else column_index(headers, spec.row_column, "row_column")
    )

    staged: list[_Staged] = []
    colors: list[str | None] = []
    # Each axis's label text -> the one value allowed to render as it.
    color_labels: dict[str, TabularCell] = {}
    group_labels: dict[str, TabularCell] = {}
    for row_index, row in enumerate(rows):
        start = _require_time(row[start_index], row_index=row_index, role="start")
        if start < 0:
            raise ValueError(
                f"the start at row {row_index} is {start}; a Gantt time must be >= 0"
            )
        if duration_index is not None:
            duration = _require_time(row[duration_index], row_index=row_index, role="duration")
            if duration < 1:
                raise ValueError(
                    f"the duration at row {row_index} is {duration}; a Gantt task must "
                    f"last at least one time unit"
                )
            end = start + duration
        else:
            assert end_index is not None  # GanttSpec requires exactly one of the two.
            end = _require_time(row[end_index], row_index=row_index, role="end")
            if end <= start:
                raise ValueError(
                    f"the end at row {row_index} is {end}, which is not after its start "
                    f"{start}; a Gantt task must last at least one time unit"
                )
        color: str | None = None
        if color_index is not None:
            raw_color: TabularCell = row[color_index]
            _check_distinct_label(
                raw_color,
                color_labels,
                placeholder=_MISSING_COLOR_NAME,
                row_index=row_index,
                role="color_column",
            )
            color = None if raw_color is None else str(raw_color)
            if color not in colors:
                colors.append(color)
        group: str | None = None
        if row_group_index is not None:
            raw_group: TabularCell = row[row_group_index]
            _check_distinct_label(
                raw_group,
                group_labels,
                placeholder=_MISSING_ROW_LABEL,
                row_index=row_index,
                role="row_column",
            )
            group = None if raw_group is None else str(raw_group)
        label = _rendered_text(row[task_index], placeholder=_MISSING_TASK_LABEL)
        staged.append(_Staged(label=label, start=start, end=end, color=color, group=group))

    if row_group_index is None:
        # One row per task, named by the task: the shape before grouping existed.
        tasks = [
            ResolvedTask(
                bar_label=None, start=task.start, end=task.end, color=task.color, grid_row=index
            )
            for index, task in enumerate(staged)
        ]
        row_labels = [task.label for task in staged]
        row_header = _TASK_COLUMN_HEADER
    else:
        tasks, row_labels = _assign_grid_rows(staged)
        row_header = spec.row_column or _TASK_COLUMN_HEADER

    horizon = max((task.end for task in tasks), default=0)
    if horizon > GANTT_MAX_HORIZON_COLUMNS:
        raise ValueError(
            f"this schedule needs a horizon of {horizon} time columns, over the "
            f"{GANTT_MAX_HORIZON_COLUMNS}-column Gantt limit; scale the times down or "
            f"write fewer tasks"
        )
    return ResolvedGantt(
        sheet_name=spec.sheet_name,
        title=spec.title,
        tasks=tuple(tasks),
        row_header=row_header,
        row_labels=tuple(row_labels),
        colors=tuple(colors),
        horizon=horizon,
    )


def _write_text(worksheet: Any, *, row: int, column: int, text: str) -> None:
    """Write literal text, never a formula.

    openpyxl infers a leading ``=`` as a formula, so every user-derived string
    this module writes forces ``data_type = "s"`` — the same rule the plain
    XLSX writer applies to the data sheet.
    """
    cell = worksheet.cell(row=row, column=column)
    cell.value = text
    cell.data_type = "s"


def render_gantt(workbook: Any, resolved: ResolvedGantt) -> str:
    """Add the Gantt sheet to ``workbook`` and return its name.

    No cell is ever merged: the read path exposes only a merge's top-left
    value, so a merge would silently lose data.
    """
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    def solid(rgb: str) -> Any:
        return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")

    # Keyed by colour value, or by None for the single default fill of a Gantt
    # that named no color_column at all.
    fills: dict[str | None, Any] = {
        color: solid(_COLOR_FILL_RGB[index % len(_COLOR_FILL_RGB)])
        for index, color in enumerate(resolved.colors)
    }
    fills.setdefault(None, solid(_COLOR_FILL_RGB[0]))

    worksheet = workbook.create_sheet(resolved.sheet_name)
    header_row = 1
    if resolved.title is not None:
        _write_text(worksheet, row=1, column=1, text=resolved.title)
        header_row = 2

    _write_text(worksheet, row=header_row, column=1, text=resolved.row_header)
    # One label per time unit plus the closing boundary, so every label reads as
    # the left edge of what follows it and the last one states the horizon: a
    # bar ending at 11 visibly stops at the tick marked 11. Only [0, horizon)
    # are unit columns; the tick can never be filled.
    for offset in range(resolved.horizon + 1):
        column = offset + 2
        worksheet.cell(row=header_row, column=column).value = offset
        worksheet.column_dimensions[get_column_letter(column)].width = _TIME_COLUMN_WIDTH

    for row_index, row_label in enumerate(resolved.row_labels):
        _write_text(worksheet, row=header_row + 1 + row_index, column=1, text=row_label)

    for task in resolved.tasks:
        row = header_row + 1 + task.grid_row
        for offset in range(task.start, task.end):
            worksheet.cell(row=row, column=offset + 2).fill = fills[task.color]
        if task.bar_label is not None:
            _write_text(worksheet, row=row, column=task.start + 2, text=task.bar_label)

    # The legend is what keeps color identity off colour alone.
    legend_row = header_row + len(resolved.row_labels) + 2
    for color_index, color in enumerate(resolved.colors):
        _write_text(
            worksheet,
            row=legend_row + color_index,
            column=1,
            text=_rendered_text(color, placeholder=_MISSING_COLOR_NAME),
        )
        worksheet.cell(row=legend_row + color_index, column=2).fill = fills[color]
    return str(worksheet.title)
