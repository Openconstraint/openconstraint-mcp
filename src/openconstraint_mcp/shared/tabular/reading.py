"""Row streaming, normalization, paging, and the byte-ceiling trim for a read."""

from __future__ import annotations

import csv
import datetime as dt
import json
import math
import sys
from collections.abc import Generator, Iterator
from contextlib import contextmanager
from pathlib import Path
from typing import Any

from ...schemas.tabular import TabularCell, TabularData
from .limits import MAX_TABULAR_RESPONSE_BYTES
from .paths import _format_for

# Python's csv module caps one field at 128 KiB and raises an opaque _csv.Error
# past it — which is not even a ValueError, so it would escape the tools' error
# translation. The response ceiling is the limit that should govern a read, so
# raise csv's cap to match it. Done once at import rather than per read: reads
# run in a thread pool, where a per-call set/restore would race (one call's
# restore would drop the cap under another call still parsing). This is an
# idempotent bump of a stdlib knob, not app state — a field past even this is
# still translated to a bounded ValueError below.
if csv.field_size_limit() < MAX_TABULAR_RESPONSE_BYTES:
    csv.field_size_limit(MAX_TABULAR_RESPONSE_BYTES)


def _normalize_header(value: object, index: int) -> str:
    """Return the string header for column ``index`` (zero-based).

    Headers are always strings. A blank position — ``None`` or an empty string
    — becomes the positional name ``col_<n>`` (one-based). Dates/times render
    ISO-8601; any other non-string renders through ``str``. Duplicate non-blank
    names are preserved: deduplicating them would be interpretation, and the
    client may legitimately have two columns of the same name.
    """
    if value is None:
        return f"col_{index + 1}"
    if isinstance(value, str):
        return value if value != "" else f"col_{index + 1}"
    if isinstance(value, dt.date | dt.time):
        return value.isoformat()
    return str(value)


def _normalize_cell(value: object) -> TabularCell:
    """Coerce one raw cell value to a JSON scalar.

    Strings, booleans, integers, and finite floats pass through with their type
    intact. Dates/times become ISO-8601 strings. Anything else — a
    ``timedelta`` from an elapsed-time cell, a stray non-finite float — renders
    through ``str`` rather than failing the read, since a value that cannot be
    a JSON scalar would otherwise make the whole page unreadable.
    """
    if value is None or isinstance(value, str | bool | int):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, dt.date | dt.time):
        return value.isoformat()
    return str(value)


def _csv_records(reader: Any, path: Path) -> Iterator[list[Any]]:
    """Yield the reader's records, translating a parse failure to a ``ValueError``.

    ``_csv.Error`` is not a ``ValueError``, so it would bypass the tools' error
    translation and surface as an opaque crash. The realistic trigger is a field
    past the size cap raised above — i.e. a field too large to return anyway.
    ``next(reader)`` also pulls more bytes from the underlying file, so a
    removable/network filesystem failing mid-read raises a raw ``OSError`` here
    too — translate that the same way.
    """
    while True:
        try:
            record = next(reader)
        except StopIteration:
            return
        except csv.Error as exc:
            raise ValueError(
                f"cannot parse {path} at line {reader.line_num}: {exc}. A single CSV field "
                f"must stay under the {MAX_TABULAR_RESPONSE_BYTES}-byte response limit."
            ) from exc
        except OSError as exc:
            raise ValueError(f"cannot read {path}: {exc}") from exc
        yield record


@contextmanager
def _open_csv_rows(path: Path) -> Generator[Iterator[list[Any]]]:
    """Yield an iterator of CSV records, using one fixed dialect.

    Comma-separated, ``"``-quoted, UTF-8 — no dialect sniffing. ``utf-8-sig``
    strips a byte-order mark if present (spreadsheet apps routinely write one),
    which is an encoding concern, not a dialect one; a plain UTF-8 file decodes
    identically.

    ``validate_tabular_read_path`` already confirms the file exists, but a
    permission-denied (or otherwise unopenable) file still raises ``OSError``
    from ``open()`` itself — not a ``ValueError`` — so it would otherwise
    bypass the tools' error translation.
    """
    try:
        handle = path.open("r", encoding="utf-8-sig", newline="")
    except OSError as exc:
        raise ValueError(f"cannot read {path}: {exc}") from exc
    try:
        yield _csv_records(csv.reader(handle), path)
    finally:
        # Guarded like the write path's staging cleanup: a close() failure
        # (e.g. a stale network-filesystem handle) must not replace whatever
        # the body above already raised, per Python's finally-exception
        # -replaces-try-exception semantics — only surface it when nothing
        # else is already propagating. The check must run BEFORE the close()
        # call: inside the `except OSError` below, sys.exc_info() always
        # reports that very close() failure as "currently being handled", so
        # checking it there would be a tautology that never fires.
        already_failing = sys.exc_info()[0] is not None
        try:
            handle.close()
        except OSError as exc:
            if not already_failing:
                raise ValueError(f"cannot read {path}: {exc}") from exc


def _xlsx_records(worksheet: Any, path: Path) -> Iterator[list[Any]]:
    """Yield ``worksheet``'s rows, translating a parse failure to a ``ValueError``.

    ``read_only=True`` streams the sheet rather than parsing it up front, so a
    corrupt worksheet's XML raises only once its rows are actually iterated —
    here, not from the ``load_workbook`` call in ``_open_xlsx_rows``. The
    raised type (``xml.etree.ElementTree.ParseError`` and others) is not a
    ``ValueError``, so it would otherwise bypass the tools' error translation.
    """
    try:
        for row in worksheet.iter_rows(values_only=True):
            yield list(row)
    except Exception as exc:
        raise ValueError(f"cannot read {path} as an XLSX workbook: {exc}") from exc


@contextmanager
def _open_xlsx_rows(
    path: Path, sheet: str | None
) -> Generator[tuple[str, list[str], Iterator[list[Any]]]]:
    """Yield ``(sheet_name, available_sheets, rows)`` for one worksheet.

    ``data_only=True`` reads a formula cell's CACHED result — the server never
    evaluates a formula, so a formula that was never calculated reads as
    ``None``. ``read_only=True`` streams the sheet rather than building the
    whole object graph, so the workbook must stay open while rows are consumed
    — hence the context manager.

    A malformed file (not a zip at all, or a zip missing the parts an XLSX
    workbook requires) raises an assortment of exception types from openpyxl's
    zip/XML layers — ``zipfile.BadZipFile``, ``KeyError``, and others — none of
    which is a ``ValueError``, so none would reach the tools' error
    translation. Translate every failure from this one call uniformly instead.

    ``workbook.sheetnames`` covers every sheet, including chart sheets — which
    have no rows and no ``iter_rows``. ``available`` is restricted to
    ``workbook.worksheets`` (data sheets only) so a chart sheet is never
    offered as a selectable name that would only fail later. The default
    selection is ``workbook.active`` when that is itself a data worksheet —
    preserving the documented "defaults to the active sheet" behavior — and
    falls back to the first data worksheet only when the active sheet is a
    chart sheet.
    """
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"cannot read {path} as an XLSX workbook: {exc}") from exc
    try:
        data_sheets = workbook.worksheets
        available = [str(ws.title) for ws in data_sheets]
        if sheet is None:
            if not data_sheets:
                if workbook.sheetnames:
                    raise ValueError(f"{path} has no worksheets (only chart sheets)")
                raise ValueError(f"{path} has no worksheets")
            active = workbook.active
            worksheet = active if active in data_sheets else data_sheets[0]
        else:
            if sheet not in available:
                if sheet in workbook.sheetnames:
                    raise ValueError(
                        f"sheet {sheet!r} in {path} is a chart sheet, not a data "
                        f"worksheet; available data sheets: {', '.join(available)}"
                    )
                names = ", ".join(available)
                raise ValueError(f"sheet {sheet!r} not found in {path}; available sheets: {names}")
            worksheet = workbook[sheet]
        yield str(worksheet.title), available, _xlsx_records(worksheet, path)
    finally:
        workbook.close()


def _row_estimate(row: list[TabularCell]) -> int:
    """Approximate the serialized byte cost of one row, plus its separating comma.

    A cheap streaming bound so a huge ``max_rows`` cannot buffer an unbounded
    page into memory. It ignores the body's fixed overhead (headers, metadata),
    which makes it a slight UNDER-estimate of the true cost — deliberate, so it
    can never stop collecting a row that would in fact have fit. The exact
    ceiling is enforced afterward against the real serialized body.
    """
    return len(json.dumps(row, ensure_ascii=False, separators=(",", ":")).encode("utf-8")) + 1


def _build_page(
    *,
    headers: list[str],
    rows: list[list[TabularCell]],
    sheet_name: str | None,
    available_sheets: list[str],
    row_offset: int,
    total_rows: int,
    byte_limited: bool,
) -> TabularData:
    """Assemble a ``TabularData`` whose pagination metadata matches its rows."""
    consumed = row_offset + len(rows)
    if consumed >= total_rows:
        return TabularData(
            headers=headers,
            rows=rows,
            sheet_name=sheet_name,
            available_sheets=available_sheets,
            row_offset=row_offset,
            next_row_offset=None,
            total_rows=total_rows,
            truncated=False,
            truncation_reason=None,
        )
    return TabularData(
        headers=headers,
        rows=rows,
        sheet_name=sheet_name,
        available_sheets=available_sheets,
        row_offset=row_offset,
        next_row_offset=consumed,
        total_rows=total_rows,
        truncated=True,
        truncation_reason="max_bytes" if byte_limited else "max_rows",
    )


def _body_size(page: TabularData) -> int:
    """Return the exact UTF-8 byte length of the page's serialized JSON body."""
    return len(page.model_dump_json().encode("utf-8"))


class _Scan:
    """What one streaming pass over the rows collected.

    ``header_values`` is the raw header record (``None`` when ``has_header`` is
    false or the file is empty), ``page`` the buffered rows for this request,
    ``total_rows`` every data row in the file, ``width`` the widest data row
    (used only to name a headerless file's columns), and ``byte_limited``
    whether buffering stopped on the byte estimate rather than ``max_rows``.
    """

    def __init__(self) -> None:
        self.header_values: list[Any] | None = None
        self.page: list[list[TabularCell]] = []
        self.total_rows: int = 0
        self.width: int = 0
        self.byte_limited: bool = False


def _scan_rows(
    raw_rows: Iterator[list[Any]],
    *,
    has_header: bool,
    row_offset: int,
    max_rows: int,
) -> _Scan:
    """Stream every row once, buffering only the requested page.

    The full pass is unavoidable — ``total_rows`` and reaching ``row_offset``
    both require it — but memory stays bounded by the page, not the file: rows
    before the offset, after ``max_rows``, and past the byte estimate are
    counted and discarded.
    """
    scan = _Scan()
    buffering = True
    estimate = 0
    for index, raw_row in enumerate(raw_rows):
        if index == 0 and has_header:
            scan.header_values = list(raw_row)
            continue
        data_index = scan.total_rows
        scan.total_rows += 1
        scan.width = max(scan.width, len(raw_row))
        if not buffering or data_index < row_offset or len(scan.page) >= max_rows:
            continue
        row = [_normalize_cell(value) for value in raw_row]
        scan.page.append(row)
        estimate += _row_estimate(row)
        if estimate > MAX_TABULAR_RESPONSE_BYTES:
            # This row already busts the ceiling, so no later row can fit
            # either. Stop buffering (but keep counting); the exact trim
            # decides whether even this row survives.
            scan.byte_limited = True
            buffering = False
    return scan


def _scan_source(
    resolved: Path,
    sheet: str | None,
    *,
    has_header: bool,
    row_offset: int,
    max_rows: int,
) -> tuple[_Scan, str | None, list[str]]:
    """Scan ``resolved`` and return ``(scan, sheet_name, available_sheets)``.

    A CSV has no sheets, so it always returns ``(scan, None, [])``.
    """
    if _format_for(resolved) == "csv":
        if sheet is not None:
            raise ValueError(f"a CSV file has no sheets, so sheet={sheet!r} cannot be selected")
        with _open_csv_rows(resolved) as csv_rows:
            scan = _scan_rows(
                csv_rows, has_header=has_header, row_offset=row_offset, max_rows=max_rows
            )
        return scan, None, []
    with _open_xlsx_rows(resolved, sheet) as (sheet_name, available_sheets, xlsx_rows):
        scan = _scan_rows(
            xlsx_rows, has_header=has_header, row_offset=row_offset, max_rows=max_rows
        )
    return scan, sheet_name, available_sheets


def _trim_to_byte_ceiling(
    *,
    headers: list[str],
    rows: list[list[TabularCell]],
    sheet_name: str | None,
    available_sheets: list[str],
    row_offset: int,
    total_rows: int,
) -> TabularData:
    """Drop trailing whole rows until the serialized body fits the ceiling.

    Only called once the full candidate is known to be over the limit, so every
    page considered here is a truncated one — a fixed metadata shape, under
    which body size is monotone in row count. That makes a binary search exact
    and bounded to ~log2(len(rows)) serializations, instead of the quadratic
    re-serialize-per-popped-row a linear scan would cost on a large page.

    Refuses rather than returning a page that makes no forward progress: a
    caller that received zero rows and the same offset back would loop forever.
    """

    def page_of(count: int) -> TabularData:
        return _build_page(
            headers=headers,
            rows=rows[:count],
            sheet_name=sheet_name,
            available_sheets=available_sheets,
            row_offset=row_offset,
            total_rows=total_rows,
            byte_limited=True,
        )

    if _body_size(page_of(0)) > MAX_TABULAR_RESPONSE_BYTES:
        raise ValueError(
            f"the normalized headers alone exceed the {MAX_TABULAR_RESPONSE_BYTES}-byte "
            f"response limit, so no page of this file can be returned"
        )

    low, high = 0, len(rows) - 1  # page_of(0) fits; the full page does not.
    while low < high:
        middle = (low + high + 1) // 2
        if _body_size(page_of(middle)) <= MAX_TABULAR_RESPONSE_BYTES:
            low = middle
        else:
            high = middle - 1
    if low == 0:
        raise ValueError(
            f"the data row at offset {row_offset} does not fit the "
            f"{MAX_TABULAR_RESPONSE_BYTES}-byte response limit on its own, so no page "
            f"starting there can make progress"
        )
    return page_of(low)
