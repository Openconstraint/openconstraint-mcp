"""The polished-table preset: resolve a ``TableStyle``, then paint a worksheet.

Split in two on purpose. ``resolve_style`` validates and computes; it never
touches a workbook, so ``core`` can run it before staging a file and a refused
style leaves the filesystem untouched. ``apply_style`` only paints what was
already resolved.

Presentation only: nothing here assigns ``.value`` or ``.data_type``, so every
XLSX round-trip guard keeps holding. A *date* ``number_format`` is the one
exception and ``resolve_style`` refuses it: openpyxl's reader converts a cell
carrying one into a ``datetime``, which ``reading`` then renders as an ISO-8601
string, so a numeric column would come back as text. Every other format code is
display-only and leaves the stored cell text alone.

Neutrals come from the ``dataviz`` skill's reference palette (chart chrome &
ink): the header fill is its hairline gridline neutral and the banded row its
page plane, both under primary ink.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from ...schemas.tabular import TableStyle, TabularCell
from .columns import column_index
from .guards import _check_no_carriage_return, _check_no_illegal_xml_characters

# Excel character units. The floor keeps a one-character column clickable; the
# ceiling stops one long cell from pushing every other column off screen.
_MIN_COLUMN_WIDTH: int = 8
_MAX_COLUMN_WIDTH: int = 60

_HEADER_FILL_RGB: str = "FFE1E0D9"
_BAND_FILL_RGB: str = "FFF9F9F7"
_HEADER_FONT_RGB: str = "FF0B0B0B"


@dataclass(frozen=True)
class ResolvedStyle:
    """One entry per column, in column order: its width and its number format."""

    widths: tuple[int, ...]
    number_formats: tuple[str | None, ...]


def _rendered_length(value: TabularCell) -> int:
    """Return the display width of one cell, counting ``None`` as blank."""
    return 0 if value is None else len(str(value))


def resolve_style(
    headers: list[str], rows: list[list[TabularCell]], style: TableStyle
) -> ResolvedStyle:
    """Resolve ``style`` against the table, or raise ``ValueError``.

    A ``columns`` key must name exactly one header: this package preserves
    duplicate headers by design, so a repeated name cannot identify a column.
    A date ``number_format`` is refused; see this module's docstring. A format
    code is a user-controlled string bound for the same XML writer as a cell,
    so it goes through the same character guards the data sheet's own values
    do.
    """
    from openpyxl.styles.numbers import is_date_format

    overrides: dict[int, Any] = {}
    for name, column in style.columns.items():
        target: int = column_index(headers, name, "style column")
        if column.number_format is not None:
            where: str = f"style column {name!r}'s number_format"
            _check_no_illegal_xml_characters(column.number_format, where)
            _check_no_carriage_return(column.number_format, where)
            if is_date_format(column.number_format):
                raise ValueError(
                    f"style column {name!r} sets the date number_format "
                    f"{column.number_format!r}; a date format makes the reader return that "
                    f"column as ISO-8601 text instead of numbers, so it is refused rather "
                    f"than silently changing the value on read-back"
                )
        overrides[target] = column

    widths: list[int] = []
    number_formats: list[str | None] = []
    for index, header in enumerate(headers):
        override = overrides.get(index)
        if override is not None and override.width is not None:
            widths.append(override.width)
        else:
            longest = max(
                [len(header), *(_rendered_length(row[index]) for row in rows)],
            )
            widths.append(min(max(longest, _MIN_COLUMN_WIDTH), _MAX_COLUMN_WIDTH))
        number_formats.append(override.number_format if override is not None else None)
    return ResolvedStyle(widths=tuple(widths), number_formats=tuple(number_formats))


def apply_style(worksheet: Any, resolved: ResolvedStyle, rows: list[list[TabularCell]]) -> None:
    """Paint the preset onto an already-populated data worksheet.

    Reads cell values nowhere: every measurement was made by ``resolve_style``.
    ``rows`` supplies only the row count the frozen pane, auto filter, and
    banding need.
    """
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(
        start_color=_HEADER_FILL_RGB, end_color=_HEADER_FILL_RGB, fill_type="solid"
    )
    band_fill = PatternFill(
        start_color=_BAND_FILL_RGB, end_color=_BAND_FILL_RGB, fill_type="solid"
    )
    header_font = Font(bold=True, color=_HEADER_FONT_RGB)

    for index, width in enumerate(resolved.widths, start=1):
        header_cell = worksheet.cell(row=1, column=index)
        header_cell.font = header_font
        header_cell.fill = header_fill
        worksheet.column_dimensions[get_column_letter(index)].width = width

    for data_index in range(len(rows)):
        for index, number_format in enumerate(resolved.number_formats, start=1):
            cell = worksheet.cell(row=data_index + 2, column=index)
            if number_format is not None:
                cell.number_format = number_format
            if data_index % 2 == 1:
                cell.fill = band_fill

    worksheet.freeze_panes = "A2"
    last_column = get_column_letter(len(resolved.widths))
    worksheet.auto_filter.ref = f"A1:{last_column}{len(rows) + 1}"
