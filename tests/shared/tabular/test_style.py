"""Tests for the polished-table preset: resolution rules and what it paints."""

from __future__ import annotations

from typing import Any

import pytest
from openpyxl import Workbook

from openconstraint_mcp.schemas.tabular import ColumnStyle, TableStyle
from openconstraint_mcp.shared.tabular.style import apply_style, resolve_style


def _worksheet(headers: list[str], rows: list[list[Any]]) -> Any:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    return worksheet


def _styled(headers: list[str], rows: list[list[Any]], style: TableStyle | None = None) -> Any:
    worksheet = _worksheet(headers, rows)
    resolved = resolve_style(headers, rows, style or TableStyle())
    apply_style(worksheet, resolved, rows)
    return worksheet


# --- resolve_style --------------------------------------------------------------


def test_resolve_rejects_a_column_key_that_is_not_a_header() -> None:
    style = TableStyle(columns={"absent": ColumnStyle(width=20)})
    with pytest.raises(ValueError, match="'absent'"):
        resolve_style(["task", "qty"], [["a", 1]], style)


def test_resolve_rejects_a_column_key_naming_a_duplicated_header() -> None:
    # Duplicate headers are preserved by design, so a name that appears twice
    # cannot identify one column.
    style = TableStyle(columns={"qty": ColumnStyle(width=20)})
    with pytest.raises(ValueError, match="ambiguous"):
        resolve_style(["qty", "qty"], [[1, 2]], style)


def test_resolve_accepts_a_duplicated_header_that_is_not_styled() -> None:
    resolved = resolve_style(["qty", "qty"], [[1, 2]], TableStyle())
    assert len(resolved.widths) == 2


def test_an_explicit_width_wins_over_the_derived_one() -> None:
    style = TableStyle(columns={"task": ColumnStyle(width=42)})
    resolved = resolve_style(["task"], [["a very long task label indeed"]], style)
    assert resolved.widths[0] == 42


def test_a_narrow_column_is_clamped_up_to_the_minimum_width() -> None:
    resolved = resolve_style(["n"], [[1]], TableStyle())
    assert resolved.widths[0] == 8


def test_a_wide_column_is_clamped_down_to_the_maximum_width() -> None:
    resolved = resolve_style(["n"], [["x" * 500]], TableStyle())
    assert resolved.widths[0] == 60


def test_a_derived_width_measures_the_widest_rendered_cell() -> None:
    resolved = resolve_style(["n"], [["x" * 20]], TableStyle())
    assert resolved.widths[0] == 20


def test_resolve_rejects_a_date_number_format() -> None:
    # openpyxl's reader turns a date-formatted cell into a datetime, which the
    # read path then renders as ISO-8601 text — a silent int -> str flip.
    style = TableStyle(columns={"day": ColumnStyle(number_format="yyyy-mm-dd")})
    with pytest.raises(ValueError, match="date number_format"):
        resolve_style(["day"], [[1]], style)


def test_resolve_accepts_a_non_date_number_format() -> None:
    style = TableStyle(columns={"qty": ColumnStyle(number_format="#,##0")})
    resolved = resolve_style(["qty"], [[1]], style)
    assert resolved.number_formats[0] == "#,##0"


@pytest.mark.parametrize("bad", ["￿", "\r"])
def test_resolve_rejects_a_number_format_xlsx_cannot_store(bad: str) -> None:
    # A format code reaches the same XML writer a cell does: unchecked, U+FFFF
    # writes a workbook no parser can reopen and a carriage return comes back
    # silently normalized.
    style = TableStyle(columns={"qty": ColumnStyle(number_format=f"0.0{bad}0")})
    with pytest.raises(ValueError, match="number_format"):
        resolve_style(["qty"], [[1]], style)


# --- apply_style ----------------------------------------------------------------


def test_the_header_row_is_bold() -> None:
    worksheet = _styled(["task", "qty"], [["a", 1]])
    assert worksheet.cell(row=1, column=1).font.bold is True


def test_the_header_row_is_filled() -> None:
    # An unstyled cell already reports fgColor.rgb == "00000000", so the fill
    # type and the preset's own colour are what distinguish styled from not.
    worksheet = _styled(["task", "qty"], [["a", 1]])
    fill = worksheet.cell(row=1, column=1).fill
    assert (fill.fill_type, fill.fgColor.rgb) == ("solid", "FFE1E0D9")


def test_the_header_row_is_frozen() -> None:
    worksheet = _styled(["task", "qty"], [["a", 1]])
    assert worksheet.freeze_panes == "A2"


def test_the_auto_filter_covers_exactly_the_used_range() -> None:
    worksheet = _styled(["task", "qty"], [["a", 1], ["b", 2]])
    assert worksheet.auto_filter.ref == "A1:B3"


def test_alternate_data_rows_are_banded() -> None:
    worksheet = _styled(["task"], [["a"], ["b"], ["c"]])
    banded = worksheet.cell(row=3, column=1).fill.fgColor.rgb
    plain = worksheet.cell(row=2, column=1).fill.fgColor.rgb
    assert banded != plain


def test_a_number_format_lands_on_the_data_cells_of_its_column() -> None:
    style = TableStyle(columns={"qty": ColumnStyle(number_format="0.00")})
    worksheet = _styled(["task", "qty"], [["a", 1.5]], style)
    assert worksheet.cell(row=2, column=2).number_format == "0.00"


def test_a_number_format_does_not_land_on_the_header_cell() -> None:
    style = TableStyle(columns={"qty": ColumnStyle(number_format="0.00")})
    worksheet = _styled(["task", "qty"], [["a", 1.5]], style)
    assert worksheet.cell(row=1, column=2).number_format != "0.00"


def test_a_column_width_reaches_the_worksheet() -> None:
    style = TableStyle(columns={"task": ColumnStyle(width=42)})
    worksheet = _styled(["task"], [["a"]], style)
    assert worksheet.column_dimensions["A"].width == 42


def test_styling_never_changes_a_cell_value() -> None:
    worksheet = _styled(["task", "qty"], [["=1+1", 3]])
    assert [worksheet.cell(row=2, column=index).value for index in (1, 2)] == ["=1+1", 3]
