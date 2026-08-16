"""Tests for the cell-grid Gantt sheet: resolution rules and what it renders."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from openconstraint_mcp.schemas.tabular import GanttSpec, TabularCell
from openconstraint_mcp.shared.tabular.gantt import render_gantt, resolve_gantt
from openconstraint_mcp.shared.tabular.limits import (
    GANTT_MAX_HORIZON_COLUMNS,
    XLSX_MAX_STRING_LENGTH,
)

_HEADERS = ["task", "start", "duration", "lane"]
_ROWS: list[list[TabularCell]] = [
    ["cut", 0, 3, "alpha"],
    ["polish", 3, 2, "beta"],
]

# The same two tasks plus a resource column, so a grouped render can share a row.
_GROUPED_HEADERS = ["task", "start", "duration", "lane", "machine"]
_GROUPED_ROWS: list[list[TabularCell]] = [
    ["cut", 0, 3, "alpha", "M1"],
    ["polish", 3, 2, "beta", "M1"],
]


def _spec(**overrides: Any) -> GanttSpec:
    fields: dict[str, Any] = {
        "task_column": "task",
        "start_column": "start",
        "duration_column": "duration",
    }
    fields.update(overrides)
    return GanttSpec(**fields)


def _rendered(
    headers: list[str] | None = None,
    rows: list[list[TabularCell]] | None = None,
    spec: GanttSpec | None = None,
) -> Any:
    workbook = Workbook()
    resolved = resolve_gantt(headers or _HEADERS, rows or _ROWS, spec or _spec())
    name = render_gantt(workbook, resolved)
    return workbook[name]


def _reloaded(worksheet: Any, tmp_path: Path) -> Any:
    """Save the rendered workbook and read it back, as a real caller would.

    An in-memory cell holding ``""`` looks fine; only the round trip shows
    openpyxl never wrote an inline string for it.
    """
    target = tmp_path / "gantt.xlsx"
    worksheet.parent.save(target)
    return load_workbook(target)[worksheet.title]


def _filled_columns(worksheet: Any, row: int, width: int) -> list[int]:
    return [
        column
        for column in range(1, width + 1)
        if worksheet.cell(row=row, column=column).fill.fill_type == "solid"
    ]


# --- resolve_gantt --------------------------------------------------------------


def test_resolve_rejects_a_column_that_is_not_a_header() -> None:
    with pytest.raises(ValueError, match="'absent'"):
        resolve_gantt(_HEADERS, _ROWS, _spec(task_column="absent"))


def test_resolve_rejects_a_duplicated_header_as_ambiguous() -> None:
    headers = ["task", "start", "start"]
    rows: list[list[TabularCell]] = [["cut", 0, 3]]
    with pytest.raises(ValueError, match="ambiguous"):
        resolve_gantt(headers, rows, _spec(duration_column=None, end_column="start"))


@pytest.mark.parametrize("start", [1.5, "0", None])
def test_resolve_rejects_a_non_integer_start(start: TabularCell) -> None:
    rows: list[list[TabularCell]] = [["cut", start, 3, "alpha"]]
    with pytest.raises(ValueError, match="row 0"):
        resolve_gantt(_HEADERS, rows, _spec())


def test_resolve_rejects_a_negative_start() -> None:
    rows: list[list[TabularCell]] = [["cut", -1, 3, "alpha"]]
    with pytest.raises(ValueError, match="row 0"):
        resolve_gantt(_HEADERS, rows, _spec())


@pytest.mark.parametrize("duration", [0, -2])
def test_resolve_rejects_a_non_positive_duration(duration: int) -> None:
    rows: list[list[TabularCell]] = [["cut", 0, duration, "alpha"]]
    with pytest.raises(ValueError, match="row 0"):
        resolve_gantt(_HEADERS, rows, _spec())


def test_resolve_rejects_an_end_before_its_start() -> None:
    headers = ["task", "start", "end"]
    rows: list[list[TabularCell]] = [["cut", 5, 2]]
    with pytest.raises(ValueError, match="row 0"):
        resolve_gantt(headers, rows, _spec(duration_column=None, end_column="end"))


def test_resolve_rejects_an_end_equal_to_its_start() -> None:
    # A zero-length task renders as a label row with no filled cells at all —
    # the same span the duration path already refuses as too short.
    headers = ["task", "start", "end"]
    rows: list[list[TabularCell]] = [["cut", 2, 2]]
    with pytest.raises(ValueError, match="row 0"):
        resolve_gantt(headers, rows, _spec(duration_column=None, end_column="end"))


def test_resolve_rejects_a_horizon_past_the_cap_and_states_it() -> None:
    over = GANTT_MAX_HORIZON_COLUMNS + 1
    rows: list[list[TabularCell]] = [["cut", 0, over, "alpha"]]
    with pytest.raises(ValueError, match=str(over)):
        resolve_gantt(_HEADERS, rows, _spec())


def test_resolve_accepts_a_horizon_exactly_at_the_cap() -> None:
    rows: list[list[TabularCell]] = [["cut", 0, GANTT_MAX_HORIZON_COLUMNS, "alpha"]]
    resolved = resolve_gantt(_HEADERS, rows, _spec())
    assert resolved.horizon == GANTT_MAX_HORIZON_COLUMNS


def test_resolve_ignores_the_title_when_checking_named_columns() -> None:
    resolved = resolve_gantt(_HEADERS, _ROWS, _spec(title="Line 1 schedule"))
    assert resolved.title == "Line 1 schedule"


@pytest.mark.parametrize("bad", ["￿", "\r"])
def test_resolve_rejects_a_sheet_name_xlsx_cannot_store(bad: str) -> None:
    # The schema enforces Excel's structural sheet-title rules; these two are
    # XML's. Unchecked, U+FFFF writes a workbook no parser can reopen and a
    # carriage return comes back silently normalized.
    with pytest.raises(ValueError, match="gantt sheet_name"):
        resolve_gantt(_HEADERS, _ROWS, _spec(sheet_name=f"Gan{bad}tt"))


@pytest.mark.parametrize("bad", ["￿", "\r"])
def test_resolve_rejects_a_title_xlsx_cannot_store(bad: str) -> None:
    with pytest.raises(ValueError, match="gantt title"):
        resolve_gantt(_HEADERS, _ROWS, _spec(title=f"Line{bad}1 schedule"))


def test_resolve_rejects_a_title_past_the_xlsx_cell_limit() -> None:
    # The title lands in a real cell, so openpyxl would truncate it silently.
    with pytest.raises(ValueError, match="gantt title"):
        resolve_gantt(_HEADERS, _ROWS, _spec(title="x" * (XLSX_MAX_STRING_LENGTH + 1)))


def test_resolve_accepts_a_title_exactly_at_the_xlsx_cell_limit() -> None:
    title = "x" * XLSX_MAX_STRING_LENGTH
    resolved = resolve_gantt(_HEADERS, _ROWS, _spec(title=title))
    assert resolved.title == title


def test_resolve_rejects_an_empty_title() -> None:
    # A "" title writes a blank A1 and still shifts the grid down a row.
    with pytest.raises(ValueError, match="gantt title"):
        resolve_gantt(_HEADERS, _ROWS, _spec(title=""))


# --- render_gantt ---------------------------------------------------------------


def test_render_returns_the_requested_sheet_name() -> None:
    workbook = Workbook()
    resolved = resolve_gantt(_HEADERS, _ROWS, _spec(sheet_name="Timeline"))
    assert render_gantt(workbook, resolved) == "Timeline"


def test_render_writes_one_row_per_task() -> None:
    worksheet = _rendered()
    assert [worksheet.cell(row=row, column=1).value for row in (2, 3)] == ["cut", "polish"]


def test_render_fills_exactly_the_columns_a_task_occupies() -> None:
    # "cut" runs [0, 3): time columns 0,1,2, i.e. sheet columns 2,3,4.
    worksheet = _rendered()
    assert _filled_columns(worksheet, row=2, width=1 + 5) == [2, 3, 4]


def test_render_offsets_a_later_task_by_its_start() -> None:
    # "polish" runs [3, 5): time columns 3,4, i.e. sheet columns 5,6.
    worksheet = _rendered()
    assert _filled_columns(worksheet, row=3, width=1 + 5) == [5, 6]


def test_render_numbers_the_time_columns_from_zero() -> None:
    worksheet = _rendered()
    assert [worksheet.cell(row=1, column=column).value for column in (2, 3, 4)] == [0, 1, 2]


def test_two_tasks_in_one_lane_share_a_fill() -> None:
    rows: list[list[TabularCell]] = [["cut", 0, 1, "alpha"], ["polish", 1, 1, "alpha"]]
    worksheet = _rendered(rows=rows, spec=_spec(lane_column="lane"))
    first = worksheet.cell(row=2, column=2).fill.fgColor.rgb
    second = worksheet.cell(row=3, column=3).fill.fgColor.rgb
    assert first == second


def test_two_lanes_get_different_fills() -> None:
    worksheet = _rendered(spec=_spec(lane_column="lane"))
    first = worksheet.cell(row=2, column=2).fill.fgColor.rgb
    second = worksheet.cell(row=3, column=5).fill.fgColor.rgb
    assert first != second


def test_a_null_lane_does_not_merge_with_a_lane_named_like_its_placeholder() -> None:
    rows: list[list[TabularCell]] = [
        ["missing", 0, 1, None],
        ["literal", 1, 1, "(no lane)"],
    ]
    worksheet: Any = _rendered(rows=rows, spec=_spec(lane_column="lane"))
    missing: str = worksheet.cell(row=2, column=2).fill.fgColor.rgb
    literal: str = worksheet.cell(row=3, column=3).fill.fgColor.rgb
    assert missing != literal


def test_a_ninth_lane_wraps_back_to_the_first_palette_slot() -> None:
    # Pins the documented cycling: the palette has eight slots, and a lane set
    # comes from caller data, so lane 9 reuses lane 1's fill. The legend is what
    # keeps identity off colour alone.
    rows: list[list[TabularCell]] = [
        [f"task{index}", index, 1, f"lane{index}"] for index in range(9)
    ]
    worksheet = _rendered(rows=rows, spec=_spec(lane_column="lane"))
    first = worksheet.cell(row=2, column=2).fill.fgColor.rgb
    ninth = worksheet.cell(row=10, column=10).fill.fgColor.rgb
    assert ninth == first


def test_a_lane_legend_names_every_lane() -> None:
    worksheet = _rendered(spec=_spec(lane_column="lane"))
    names = [
        cell.value
        for row in worksheet.iter_rows(min_row=4, min_col=1, max_col=1)
        for cell in row
        if cell.value is not None
    ]
    assert names == ["alpha", "beta"]


def test_render_merges_no_cells() -> None:
    # The read path exposes only a merge's top-left value, so a merge would
    # silently lose data.
    worksheet = _rendered(spec=_spec(lane_column="lane"))
    assert list(worksheet.merged_cells.ranges) == []


def test_a_formula_looking_task_label_is_written_as_a_string_cell() -> None:
    rows: list[list[TabularCell]] = [["=HYPERLINK(1)", 0, 1, "alpha"]]
    worksheet = _rendered(rows=rows)
    assert worksheet.cell(row=2, column=1).data_type == "s"


def test_a_formula_looking_lane_name_is_written_as_a_string_cell() -> None:
    # One task: header row 1, the task row 2, the legend two rows below it.
    rows: list[list[TabularCell]] = [["cut", 0, 1, "=1+1"]]
    worksheet = _rendered(rows=rows, spec=_spec(lane_column="lane"))
    assert worksheet.cell(row=4, column=1).data_type == "s"


def test_a_null_task_label_reads_back_as_a_placeholder(tmp_path: Path) -> None:
    rows: list[list[TabularCell]] = [[None, 0, 1, "alpha"]]
    worksheet = _reloaded(_rendered(rows=rows), tmp_path)
    assert worksheet.cell(row=2, column=1).value == "(untitled task)"


def test_a_null_lane_reads_back_as_a_named_legend_entry(tmp_path: Path) -> None:
    # A blank name beside a coloured swatch would put lane identity on colour alone.
    rows: list[list[TabularCell]] = [["cut", 0, 1, None]]
    worksheet = _reloaded(_rendered(rows=rows, spec=_spec(lane_column="lane")), tmp_path)
    assert worksheet.cell(row=4, column=1).value == "(no lane)"


def test_a_title_lands_in_the_first_cell() -> None:
    worksheet = _rendered(spec=_spec(title="Line 1 schedule"))
    assert worksheet.cell(row=1, column=1).value == "Line 1 schedule"


def test_a_formula_looking_title_is_written_as_a_string_cell() -> None:
    worksheet = _rendered(spec=_spec(title="=1+1"))
    assert worksheet.cell(row=1, column=1).data_type == "s"


def test_a_title_shifts_the_grid_down_one_row() -> None:
    worksheet = _rendered(spec=_spec(title="Line 1 schedule"))
    assert [worksheet.cell(row=row, column=1).value for row in (3, 4)] == ["cut", "polish"]


# --- the closing axis tick ------------------------------------------------------


def test_the_time_axis_ends_with_a_closing_tick() -> None:
    # _ROWS spans [0, 5): five unit columns labelled 0..4 in sheet columns 2..6,
    # then the boundary the last bar ends at, so a makespan reads off the axis.
    worksheet = _rendered()
    assert worksheet.cell(row=1, column=7).value == 5


def test_the_closing_tick_column_is_never_filled() -> None:
    # It marks a boundary, not a sixth time unit; a bar can only occupy [0, horizon).
    worksheet = _rendered()
    assert _filled_columns(worksheet, row=2, width=7) == [2, 3, 4]


# --- row_column grouping --------------------------------------------------------


def test_resolve_rejects_a_row_column_that_is_not_a_header() -> None:
    with pytest.raises(ValueError, match="'absent'"):
        resolve_gantt(_GROUPED_HEADERS, _GROUPED_ROWS, _spec(row_column="absent"))


def test_resolve_rejects_a_row_column_naming_a_duplicated_header() -> None:
    headers = ["task", "start", "duration", "machine", "machine"]
    rows: list[list[TabularCell]] = [["cut", 0, 3, "M1", "M1"]]
    with pytest.raises(ValueError, match="ambiguous"):
        resolve_gantt(headers, rows, _spec(row_column="machine"))


def test_tasks_sharing_a_row_value_share_one_grid_row() -> None:
    worksheet = _rendered(
        headers=_GROUPED_HEADERS, rows=_GROUPED_ROWS, spec=_spec(row_column="machine")
    )
    # "cut" [0,3) and "polish" [3,5) both land on the single M1 row.
    assert _filled_columns(worksheet, row=2, width=7) == [2, 3, 4, 5, 6]


def test_a_grouped_row_is_named_by_its_row_column_value() -> None:
    worksheet = _rendered(
        headers=_GROUPED_HEADERS, rows=_GROUPED_ROWS, spec=_spec(row_column="machine")
    )
    assert worksheet.cell(row=2, column=1).value == "M1"


def test_a_grouped_grid_heads_its_first_column_with_the_row_columns_header() -> None:
    worksheet = _rendered(
        headers=_GROUPED_HEADERS, rows=_GROUPED_ROWS, spec=_spec(row_column="machine")
    )
    assert worksheet.cell(row=1, column=1).value == "machine"


def test_an_ungrouped_grid_still_heads_its_first_column_with_task() -> None:
    assert _rendered().cell(row=1, column=1).value == "Task"


def test_a_grouped_bar_carries_its_task_label() -> None:
    # Column A now names the resource, so task identity moves into the bar.
    worksheet = _rendered(
        headers=_GROUPED_HEADERS, rows=_GROUPED_ROWS, spec=_spec(row_column="machine")
    )
    assert [worksheet.cell(row=2, column=column).value for column in (2, 5)] == ["cut", "polish"]


def test_an_ungrouped_bar_carries_no_label() -> None:
    # Column A already names the task, so a bar label would only duplicate it.
    worksheet = _rendered()
    assert worksheet.cell(row=2, column=2).value is None


def test_a_formula_looking_bar_label_is_written_as_a_string_cell() -> None:
    rows: list[list[TabularCell]] = [["=1+1", 0, 1, "alpha", "M1"]]
    worksheet = _rendered(headers=_GROUPED_HEADERS, rows=rows, spec=_spec(row_column="machine"))
    assert worksheet.cell(row=2, column=2).data_type == "s"


def test_overlapping_tasks_in_one_group_spill_onto_sub_rows() -> None:
    # One resource, but the two spans collide, so M1 cannot be a single row.
    rows: list[list[TabularCell]] = [
        ["cut", 0, 3, "alpha", "M1"],
        ["polish", 1, 3, "beta", "M1"],
    ]
    resolved = resolve_gantt(_GROUPED_HEADERS, rows, _spec(row_column="machine"))
    assert resolved.row_labels == ("M1", "M1")


def test_a_spilled_sub_row_repeats_its_row_name() -> None:
    # Identity must never rest on position: every sub-row names its resource.
    rows: list[list[TabularCell]] = [
        ["cut", 0, 3, "alpha", "M1"],
        ["polish", 1, 3, "beta", "M1"],
    ]
    worksheet = _rendered(headers=_GROUPED_HEADERS, rows=rows, spec=_spec(row_column="machine"))
    assert [worksheet.cell(row=row, column=1).value for row in (2, 3)] == ["M1", "M1"]


def test_a_freed_sub_row_is_reused_by_a_later_task() -> None:
    # a[0,2) and b[2,4) fit one sub-row; only c[0,4) needs a second.
    rows: list[list[TabularCell]] = [
        ["a", 0, 2, "alpha", "M1"],
        ["b", 2, 2, "alpha", "M1"],
        ["c", 0, 4, "beta", "M1"],
    ]
    worksheet = _rendered(headers=_GROUPED_HEADERS, rows=rows, spec=_spec(row_column="machine"))
    assert [worksheet.cell(row=2, column=column).value for column in (2, 4)] == ["a", "b"]


def test_a_three_deep_overlap_opens_three_sub_rows() -> None:
    rows: list[list[TabularCell]] = [
        [name, 0, 3, "alpha", "M1"] for name in ("a", "b", "c")
    ]
    worksheet = _rendered(headers=_GROUPED_HEADERS, rows=rows, spec=_spec(row_column="machine"))
    assert [worksheet.cell(row=row, column=2).value for row in (2, 3, 4)] == ["a", "b", "c"]


def test_a_null_row_value_reads_back_as_a_placeholder(tmp_path: Path) -> None:
    rows: list[list[TabularCell]] = [["cut", 0, 1, "alpha", None]]
    worksheet = _reloaded(
        _rendered(headers=_GROUPED_HEADERS, rows=rows, spec=_spec(row_column="machine")),
        tmp_path,
    )
    assert worksheet.cell(row=2, column=1).value == "(no group)"


def test_grouping_puts_the_lane_legend_below_the_collapsed_grid() -> None:
    # Two tasks collapse to one M1 row, so the legend rises a row with them.
    worksheet = _rendered(
        headers=_GROUPED_HEADERS,
        rows=_GROUPED_ROWS,
        spec=_spec(row_column="machine", lane_column="lane"),
    )
    assert [worksheet.cell(row=row, column=1).value for row in (4, 5)] == ["alpha", "beta"]
