"""Tests for the plain table emitters: what the CSV and XLSX writers put on disk."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from openconstraint_mcp.shared.tabular.core import read_tabular_data, write_tabular_data


def test_csv_write_round_trips_through_the_reader(tmp_path: Path) -> None:
    target = tmp_path / "out.csv"
    write_tabular_data(["name", "qty"], [["widget", "3"]], target)
    page = read_tabular_data(target)
    assert (page.headers, page.rows) == (["name", "qty"], [["widget", "3"]])


def test_csv_write_emits_none_as_an_empty_field(tmp_path: Path) -> None:
    target = tmp_path / "out.csv"
    write_tabular_data(["a", "b"], [["x", None]], target)
    assert target.read_text(encoding="utf-8").splitlines()[1] == "x,"


def test_xlsx_write_round_trips_scalar_types(tmp_path: Path) -> None:
    target = tmp_path / "out.xlsx"
    write_tabular_data(["s", "i", "f", "b", "n"], [["x", 7, 1.5, True, None]], target)
    assert read_tabular_data(target).rows == [["x", 7, 1.5, True, None]]


def test_xlsx_write_names_the_sheet_sheet1(tmp_path: Path) -> None:
    target = tmp_path / "out.xlsx"
    write_tabular_data(["a"], [["1"]], target)
    assert read_tabular_data(target).sheet_name == "Sheet1"


def test_xlsx_write_stores_a_formula_looking_string_as_literal_text(tmp_path: Path) -> None:
    target = tmp_path / "out.xlsx"
    write_tabular_data(["expr"], [["=1+1"]], target)
    assert read_tabular_data(target).rows == [["=1+1"]]


def test_xlsx_write_marks_a_formula_looking_string_as_a_string_cell(tmp_path: Path) -> None:
    # The stronger claim: not merely that it reads back as text, but that the
    # cell is typed as a string, so Excel/Calc will not evaluate it either.
    target = tmp_path / "out.xlsx"
    write_tabular_data(["expr"], [["=1+1"]], target)
    workbook = load_workbook(target, data_only=False)
    assert workbook["Sheet1"].cell(row=2, column=1).data_type == "s"
