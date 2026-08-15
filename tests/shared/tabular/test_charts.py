"""Tests for the chart sheets: resolution rules and what gets anchored where."""

from __future__ import annotations

from typing import Any

import pytest
from openpyxl import Workbook

from openconstraint_mcp.schemas.tabular import ChartSpec, TabularCell
from openconstraint_mcp.shared.tabular.charts import render_charts, resolve_charts
from openconstraint_mcp.shared.tabular.limits import WRITE_SHEET_NAME

_HEADERS = ["task", "qty", "cost"]
_ROWS: list[list[TabularCell]] = [["cut", 3, 1.5], ["polish", 10, 2.5]]


def _data_workbook() -> Any:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = WRITE_SHEET_NAME
    worksheet.append(_HEADERS)
    for row in _ROWS:
        worksheet.append(row)
    return workbook


def _render(specs: list[ChartSpec], rows: list[list[TabularCell]] | None = None) -> Any:
    workbook = _data_workbook()
    resolved = resolve_charts(_HEADERS, rows or _ROWS, specs)
    names = render_charts(workbook, resolved)
    return workbook, names


def _spec(**overrides: Any) -> ChartSpec:
    fields: dict[str, Any] = {"kind": "bar", "x_column": "task", "y_columns": ["qty"]}
    fields.update(overrides)
    return ChartSpec(**fields)


# --- resolve_charts -------------------------------------------------------------


def test_resolve_rejects_an_x_column_that_is_not_a_header() -> None:
    with pytest.raises(ValueError, match="'absent'"):
        resolve_charts(_HEADERS, _ROWS, [_spec(x_column="absent")])


def test_resolve_rejects_a_y_column_that_is_not_a_header() -> None:
    with pytest.raises(ValueError, match="'absent'"):
        resolve_charts(_HEADERS, _ROWS, [_spec(y_columns=["qty", "absent"])])


def test_resolve_rejects_a_duplicated_header_as_ambiguous() -> None:
    headers = ["task", "qty", "qty"]
    rows: list[list[TabularCell]] = [["cut", 3, 4]]
    with pytest.raises(ValueError, match="ambiguous"):
        resolve_charts(headers, rows, [_spec()])


@pytest.mark.parametrize("value", ["3", None, True])
def test_resolve_rejects_a_non_numeric_y_value(value: TabularCell) -> None:
    rows: list[list[TabularCell]] = [["cut", value, 1.5]]
    with pytest.raises(ValueError, match="row 0"):
        resolve_charts(_HEADERS, rows, [_spec()])


def test_resolve_rejects_a_non_numeric_x_value_for_a_scatter() -> None:
    # ScatterChart's x_axis is a NumericAxis, unlike bar/line's category axis.
    with pytest.raises(ValueError, match="row 0"):
        resolve_charts(_HEADERS, _ROWS, [_spec(kind="scatter", x_column="task")])


def test_resolve_accepts_a_non_numeric_x_value_for_a_bar() -> None:
    resolved = resolve_charts(_HEADERS, _ROWS, [_spec()])
    assert len(resolved) == 1


def test_resolve_rejects_two_sheet_names_differing_only_by_case() -> None:
    specs = [_spec(sheet_name="Charts"), _spec(sheet_name="charts")]
    with pytest.raises(ValueError, match="ambiguous"):
        resolve_charts(_HEADERS, _ROWS, specs)


@pytest.mark.parametrize("bad", ["￿", "\r"])
def test_resolve_rejects_a_sheet_name_xlsx_cannot_store(bad: str) -> None:
    # The schema enforces Excel's structural sheet-title rules; these two are
    # XML's. Unchecked, U+FFFF writes a workbook no parser can reopen and a
    # carriage return comes back silently normalized.
    with pytest.raises(ValueError, match="chart sheet_name"):
        resolve_charts(_HEADERS, _ROWS, [_spec(sheet_name=f"Char{bad}ts")])


@pytest.mark.parametrize("bad", ["￿", "\r"])
def test_resolve_rejects_a_title_xlsx_cannot_store(bad: str) -> None:
    with pytest.raises(ValueError, match="chart title"):
        resolve_charts(_HEADERS, _ROWS, [_spec(title=f"Quantity{bad}by task")])


# --- render_charts --------------------------------------------------------------


def test_render_returns_one_sheet_name_per_distinct_sheet() -> None:
    specs = [_spec(sheet_name="Charts"), _spec(kind="line", sheet_name="Trends")]
    _, names = _render(specs)
    assert names == ["Charts", "Trends"]


def test_two_specs_sharing_a_sheet_name_land_on_one_sheet() -> None:
    specs = [_spec(sheet_name="Charts"), _spec(kind="line", sheet_name="Charts")]
    workbook, names = _render(specs)
    assert (names, len(workbook["Charts"]._charts)) == (["Charts"], 2)


def test_a_charts_series_count_matches_its_y_columns() -> None:
    workbook, _ = _render([_spec(y_columns=["qty", "cost"])])
    assert len(workbook["Charts"]._charts[0].series) == 2


def test_a_scatter_charts_series_count_matches_its_y_columns() -> None:
    workbook, _ = _render([_spec(kind="scatter", x_column="qty", y_columns=["cost"])])
    assert len(workbook["Charts"]._charts[0].series) == 1


def test_a_chart_title_becomes_the_chart_objects_own_title() -> None:
    workbook, _ = _render([_spec(title="Quantity by task")])
    chart = workbook["Charts"]._charts[0]
    assert chart.title.tx.rich.p[0].r[0].t == "Quantity by task"


def test_an_untitled_chart_has_no_title() -> None:
    workbook, _ = _render([_spec()])
    assert workbook["Charts"]._charts[0].title is None


def test_rendering_does_not_modify_the_data_sheet() -> None:
    workbook, _ = _render([_spec(y_columns=["qty", "cost"])])
    worksheet = workbook[WRITE_SHEET_NAME]
    values = [[cell.value for cell in row] for row in worksheet.iter_rows()]
    assert values == [_HEADERS, ["cut", 3, 1.5], ["polish", 10, 2.5]]


def test_a_chart_sheet_holds_no_rows() -> None:
    workbook, _ = _render([_spec()])
    assert list(workbook["Charts"].iter_rows()) == []
