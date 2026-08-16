"""Tests for the bounded read path: streaming, normalization, paging, trimming.

These use real files under ``tmp_path``: the leaf's whole job is filesystem
behavior, so mocking the filesystem would test nothing.
"""

from __future__ import annotations

import csv
import datetime as dt
import os
import re
import sys
import zipfile
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from openpyxl import Workbook

from openconstraint_mcp.shared.tabular.core import read_tabular_data
from openconstraint_mcp.shared.tabular.limits import MAX_TABULAR_RESPONSE_BYTES

_LIBREOFFICE_FIXTURE = (
    Path(__file__).parent.parent.parent / "fixtures" / "libreoffice_inventory.xlsx"
)


def _write_csv(path: Path, records: list[list[object]]) -> Path:
    with path.open("w", encoding="utf-8", newline="") as handle:
        csv.writer(handle).writerows(records)
    return path


def _write_xlsx(path: Path, records: list[list[object]], *, sheets: dict | None = None) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    for record in records:
        worksheet.append(record)
    for name, rows in (sheets or {}).items():
        extra = workbook.create_sheet(name)
        for record in rows:
            extra.append(record)
    workbook.save(path)
    return path


# --- reading: CSV ----------------------------------------------------------------


def test_csv_read_returns_headers_and_rows(tmp_path: Path) -> None:
    path = _write_csv(tmp_path / "d.csv", [["name", "qty"], ["widget", "3"], ["gadget", "10"]])
    page = read_tabular_data(path)
    assert (page.headers, page.rows) == (["name", "qty"], [["widget", "3"], ["gadget", "10"]])


def test_csv_read_keeps_every_cell_textual(tmp_path: Path) -> None:
    # CSV carries no types: a numeric-looking field stays the string it was.
    path = _write_csv(tmp_path / "d.csv", [["qty"], ["3"]])
    assert read_tabular_data(path).rows == [["3"]]


def test_csv_read_reports_no_sheets(tmp_path: Path) -> None:
    path = _write_csv(tmp_path / "d.csv", [["a"], ["1"]])
    page = read_tabular_data(path)
    assert (page.sheet_name, page.available_sheets) == (None, [])


def test_csv_read_rejects_a_sheet_selection(tmp_path: Path) -> None:
    path = _write_csv(tmp_path / "d.csv", [["a"], ["1"]])
    with pytest.raises(ValueError, match="no sheets"):
        read_tabular_data(path, sheet="Sheet1")


def test_csv_read_strips_a_byte_order_mark_from_the_first_header(tmp_path: Path) -> None:
    path = tmp_path / "bom.csv"
    path.write_bytes("﻿name,qty\r\nwidget,3\r\n".encode())
    assert read_tabular_data(path).headers == ["name", "qty"]


def test_headerless_csv_read_names_columns_positionally(tmp_path: Path) -> None:
    path = _write_csv(tmp_path / "d.csv", [["widget", "3"], ["gadget", "10"]])
    page = read_tabular_data(path, has_header=False)
    assert (page.headers, page.total_rows) == (["col_1", "col_2"], 2)


def test_headerless_positional_names_cover_the_widest_row_in_the_file(tmp_path: Path) -> None:
    # Headers must be stable across pages, so they are derived from the whole
    # file — not from whichever rows this page happens to hold.
    path = _write_csv(tmp_path / "d.csv", [["a"], ["b", "c", "d"]])
    page = read_tabular_data(path, has_header=False, max_rows=1)
    assert page.headers == ["col_1", "col_2", "col_3"]


def test_a_data_row_wider_than_the_header_row_gets_positional_names_for_the_extra_cells(
    tmp_path: Path,
) -> None:
    path = _write_csv(tmp_path / "d.csv", [["name"], ["widget", "3"]])
    page = read_tabular_data(path)
    assert (page.headers, page.rows) == (["name", "col_2"], [["widget", "3"]])


def test_empty_csv_returns_no_headers_and_no_rows(tmp_path: Path) -> None:
    path = _write_csv(tmp_path / "d.csv", [])
    page = read_tabular_data(path)
    assert (page.headers, page.rows, page.total_rows) == ([], [], 0)


def test_header_only_csv_keeps_its_headers(tmp_path: Path) -> None:
    path = _write_csv(tmp_path / "d.csv", [["name", "qty"]])
    page = read_tabular_data(path)
    assert (page.headers, page.rows, page.total_rows) == (["name", "qty"], [], 0)


# --- reading: XLSX ---------------------------------------------------------------


def test_xlsx_read_returns_headers_and_rows(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [["name", "qty"], ["widget", 3]])
    page = read_tabular_data(path)
    assert (page.headers, page.rows) == (["name", "qty"], [["widget", 3]])


def test_xlsx_read_preserves_numeric_and_boolean_cell_types(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [["i", "f", "b"], [7, 1.5, True]])
    assert read_tabular_data(path).rows == [[7, 1.5, True]]


def test_xlsx_read_converts_a_datetime_cell_to_iso_8601(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [["due"], [dt.datetime(2026, 7, 12, 9, 30)]])
    assert read_tabular_data(path).rows == [["2026-07-12T09:30:00"]]


def test_xlsx_read_defaults_to_the_active_sheet(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [["a"], [1]], sheets={"Other": [["b"], [2]]})
    page = read_tabular_data(path)
    assert (page.sheet_name, page.rows) == ("Sheet1", [[1]])


def test_xlsx_read_defaults_to_the_active_sheet_even_when_it_is_not_first(
    tmp_path: Path,
) -> None:
    # The active sheet need not be the first one in the workbook; the default
    # selection must follow workbook.active, not just take index 0.
    path = tmp_path / "d.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["a"])
    first.append([1])
    other = workbook.create_sheet("Other")
    other.append(["b"])
    other.append([2])
    workbook.active = other
    workbook.save(path)
    page = read_tabular_data(path)
    assert (page.sheet_name, page.rows) == ("Other", [[2]])


def test_xlsx_read_selects_an_explicit_sheet(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [["a"], [1]], sheets={"Other": [["b"], [2]]})
    page = read_tabular_data(path, sheet="Other")
    assert (page.sheet_name, page.rows) == ("Other", [[2]])


def test_xlsx_read_reports_every_available_sheet(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [["a"], [1]], sheets={"Other": [["b"], [2]]})
    assert read_tabular_data(path).available_sheets == ["Sheet1", "Other"]


def test_xlsx_read_rejects_an_unknown_sheet_and_names_the_available_ones(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [["a"], [1]])
    with pytest.raises(ValueError, match="not found.*available sheets: Sheet1"):
        read_tabular_data(path, sheet="Missing")


def _chartsheet_workbook(*, active_is_chart: bool) -> Workbook:
    # A real chartsheet-containing file can't round-trip through this
    # openpyxl version's writer/reader (an unrelated openpyxl bug in its
    # chartsheet relationship handling), so the workbook is built in memory
    # and substituted for ``openpyxl.load_workbook``'s return value instead.
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data.append(["a"])
    data.append([1])
    chart = workbook.create_chartsheet(title="Chart1")
    workbook.active = chart if active_is_chart else data
    return workbook


def test_xlsx_read_skips_a_chart_sheet_when_selecting_the_default(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [["a"], [1]])
    workbook = _chartsheet_workbook(active_is_chart=True)
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *a, **k: workbook)
    page = read_tabular_data(path)
    assert (page.sheet_name, page.rows) == ("Data", [[1]])


def test_xlsx_read_excludes_chart_sheets_from_available_sheets(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [["a"], [1]])
    workbook = _chartsheet_workbook(active_is_chart=False)
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *a, **k: workbook)
    assert read_tabular_data(path).available_sheets == ["Data"]


def test_xlsx_read_rejects_selecting_a_chart_sheet_by_name(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [["a"], [1]])
    workbook = _chartsheet_workbook(active_is_chart=False)
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *a, **k: workbook)
    with pytest.raises(ValueError, match="'Chart1'.*chart sheet, not a data worksheet"):
        read_tabular_data(path, sheet="Chart1")


def test_xlsx_read_reports_a_chart_sheets_only_workbook_as_having_no_worksheets(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [["a"], [1]])
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_chartsheet(title="Chart1")
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *a, **k: workbook)
    with pytest.raises(ValueError, match="no worksheets \\(only chart sheets\\)"):
        read_tabular_data(path)


def test_empty_xlsx_returns_no_headers_and_no_rows(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [])
    page = read_tabular_data(path)
    assert (page.headers, page.rows, page.total_rows) == ([], [], 0)


def test_header_only_xlsx_keeps_its_headers(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [["name", "qty"]])
    page = read_tabular_data(path)
    assert (page.headers, page.rows, page.total_rows) == (["name", "qty"], [], 0)


# --- reading: header normalization ------------------------------------------------


def test_numeric_header_becomes_a_string(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [[2026, "name"], [1, "widget"]])
    assert read_tabular_data(path).headers == ["2026", "name"]


def test_boolean_header_becomes_a_string(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [[True, "name"], [1, "widget"]])
    assert read_tabular_data(path).headers == ["True", "name"]


def test_datetime_header_becomes_an_iso_8601_string(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [[dt.datetime(2026, 7, 12), "n"], [1, "w"]])
    assert read_tabular_data(path).headers == ["2026-07-12T00:00:00", "n"]


def test_blank_xlsx_header_becomes_a_positional_name(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path / "d.xlsx", [["name", None, "qty"], ["widget", "x", 3]])
    assert read_tabular_data(path).headers == ["name", "col_2", "qty"]


def test_empty_string_header_becomes_a_positional_name(tmp_path: Path) -> None:
    path = _write_csv(tmp_path / "d.csv", [["name", "", "qty"], ["widget", "x", "3"]])
    assert read_tabular_data(path).headers == ["name", "col_2", "qty"]


def test_duplicate_headers_are_preserved(tmp_path: Path) -> None:
    path = _write_csv(tmp_path / "d.csv", [["qty", "qty"], ["1", "2"]])
    assert read_tabular_data(path).headers == ["qty", "qty"]


# --- reading: a workbook written by a non-openpyxl writer --------------------------


def test_libreoffice_written_workbook_is_a_genuine_non_openpyxl_fixture() -> None:
    # Guards the point of the fixture: if it were ever regenerated by openpyxl,
    # the read test below would silently stop proving anything about real-world
    # XLSX files.
    with zipfile.ZipFile(_LIBREOFFICE_FIXTURE) as archive:
        app_xml = archive.read("docProps/app.xml").decode("utf-8")
    assert "LibreOffice" in app_xml


def test_libreoffice_written_workbook_reads_with_correct_types() -> None:
    page = read_tabular_data(_LIBREOFFICE_FIXTURE)
    assert page.headers == ["name", "qty", "price", "due"]
    assert page.rows == [
        ["widget", 3, 4.5, "2026-07-12T00:00:00"],
        ["gadget", 10, 19.99, "2026-08-01T00:00:00"],
    ]


def test_libreoffice_written_workbook_reports_its_sheet_name() -> None:
    page = read_tabular_data(_LIBREOFFICE_FIXTURE)
    assert (page.sheet_name, page.available_sheets) == ("inventory", ["inventory"])


# --- reading: pagination -----------------------------------------------------------


def _numbered_csv(path: Path, count: int) -> Path:
    return _write_csv(path, [["n"], *[[str(i)] for i in range(count)]])


def test_row_limit_truncates_and_reports_the_next_offset(tmp_path: Path) -> None:
    path = _numbered_csv(tmp_path / "d.csv", 5)
    page = read_tabular_data(path, max_rows=2)
    assert (page.rows, page.next_row_offset, page.truncated, page.truncation_reason) == (
        [["0"], ["1"]],
        2,
        True,
        "max_rows",
    )


def test_row_limit_page_reports_the_files_total_row_count(tmp_path: Path) -> None:
    path = _numbered_csv(tmp_path / "d.csv", 5)
    assert read_tabular_data(path, max_rows=2).total_rows == 5


def test_next_row_offset_reaches_the_following_page(tmp_path: Path) -> None:
    path = _numbered_csv(tmp_path / "d.csv", 5)
    first = read_tabular_data(path, max_rows=2)
    assert first.next_row_offset is not None
    second = read_tabular_data(path, row_offset=first.next_row_offset, max_rows=2)
    assert (second.rows, second.row_offset, second.next_row_offset) == ([["2"], ["3"]], 2, 4)


def test_the_final_page_is_not_truncated(tmp_path: Path) -> None:
    path = _numbered_csv(tmp_path / "d.csv", 5)
    page = read_tabular_data(path, row_offset=4, max_rows=2)
    assert (page.rows, page.truncated, page.next_row_offset, page.truncation_reason) == (
        [["4"]],
        False,
        None,
        None,
    )


def test_an_offset_past_the_end_returns_an_empty_untruncated_page(tmp_path: Path) -> None:
    path = _numbered_csv(tmp_path / "d.csv", 3)
    page = read_tabular_data(path, row_offset=99)
    assert (page.rows, page.truncated, page.total_rows, page.headers) == ([], False, 3, ["n"])


def test_a_page_repeats_the_headers_so_it_is_self_describing(tmp_path: Path) -> None:
    path = _numbered_csv(tmp_path / "d.csv", 5)
    assert read_tabular_data(path, row_offset=3, max_rows=1).headers == ["n"]


def test_negative_row_offset_is_rejected(tmp_path: Path) -> None:
    path = _numbered_csv(tmp_path / "d.csv", 3)
    with pytest.raises(ValueError, match="row_offset must be >= 0"):
        read_tabular_data(path, row_offset=-1)


def test_non_positive_max_rows_is_rejected(tmp_path: Path) -> None:
    path = _numbered_csv(tmp_path / "d.csv", 3)
    with pytest.raises(ValueError, match="max_rows must be >= 1"):
        read_tabular_data(path, max_rows=0)


# --- reading: the byte ceiling -----------------------------------------------------


def _fat_row_csv(tmp_path: Path, *, rows: int, cell_bytes: int) -> Path:
    payload = "x" * cell_bytes
    return _write_csv(tmp_path / "fat.csv", [["blob"], *[[payload] for _ in range(rows)]])


def test_byte_ceiling_truncates_a_page_that_would_exceed_it(tmp_path: Path) -> None:
    # 20 rows x ~100 KiB each = ~2 MiB, well past the 1 MiB body ceiling, and
    # max_rows would happily allow all 20.
    path = _fat_row_csv(tmp_path, rows=20, cell_bytes=100_000)
    page = read_tabular_data(path, max_rows=20)
    assert page.truncated and page.truncation_reason == "max_bytes"


def test_a_byte_truncated_page_stays_within_the_ceiling(tmp_path: Path) -> None:
    path = _fat_row_csv(tmp_path, rows=20, cell_bytes=100_000)
    page = read_tabular_data(path, max_rows=20)
    assert len(page.model_dump_json().encode("utf-8")) <= MAX_TABULAR_RESPONSE_BYTES


def test_a_byte_truncated_page_returns_only_whole_rows(tmp_path: Path) -> None:
    payload = "x" * 100_000
    path = _fat_row_csv(tmp_path, rows=20, cell_bytes=100_000)
    page = read_tabular_data(path, max_rows=20)
    assert all(row == [payload] for row in page.rows)


def test_a_byte_truncated_page_resumes_exactly_where_it_stopped(tmp_path: Path) -> None:
    path = _fat_row_csv(tmp_path, rows=20, cell_bytes=100_000)
    page = read_tabular_data(path, max_rows=20)
    assert page.next_row_offset == len(page.rows)


def test_a_page_that_exactly_fits_the_ceiling_is_returned_whole(tmp_path: Path) -> None:
    # Grow a payload until the body lands as close under the ceiling as one
    # byte of cell content allows, then assert nothing was trimmed. This is the
    # regression guard for an over-eager estimator trimming a page that fits.
    path = tmp_path / "exact.csv"
    size = 900_000
    while True:
        _write_csv(path, [["blob"], ["x" * size]])
        body = len(read_tabular_data(path).model_dump_json().encode("utf-8"))
        if body >= MAX_TABULAR_RESPONSE_BYTES - 1:
            break
        size += MAX_TABULAR_RESPONSE_BYTES - body
    page = read_tabular_data(path)
    assert len(page.model_dump_json().encode("utf-8")) <= MAX_TABULAR_RESPONSE_BYTES
    assert (page.rows, page.truncated) == ([["x" * size]], False)


def _oversized_record(cells: int = 40_000) -> list[object]:
    """A single record too wide to fit the ceiling, without any one huge field."""
    return ["x" * 30] * cells


def test_a_single_row_larger_than_the_ceiling_is_a_bounded_error(tmp_path: Path) -> None:
    path = _write_csv(tmp_path / "d.csv", [["blob"], _oversized_record()])
    with pytest.raises(ValueError, match="data row at offset 0 does not fit"):
        read_tabular_data(path)


def test_an_oversized_row_error_names_the_offending_offset(tmp_path: Path) -> None:
    path = _write_csv(tmp_path / "d.csv", [["blob"], ["small"], _oversized_record()])
    with pytest.raises(ValueError, match="data row at offset 1 does not fit"):
        read_tabular_data(path, row_offset=1)


def test_headers_larger_than_the_ceiling_are_a_bounded_error(tmp_path: Path) -> None:
    path = _write_csv(tmp_path / "d.csv", [_oversized_record(), ["1"]])
    with pytest.raises(ValueError, match="headers alone exceed"):
        read_tabular_data(path)


def test_a_csv_field_past_the_field_limit_is_a_bounded_value_error(tmp_path: Path) -> None:
    # Python's csv module raises _csv.Error — not a ValueError — for an
    # oversized field, which would otherwise escape the tools' error handling.
    path = _write_csv(tmp_path / "d.csv", [["blob"], ["x" * (MAX_TABULAR_RESPONSE_BYTES + 10)]])
    with pytest.raises(ValueError, match="A single CSV field must stay under"):
        read_tabular_data(path)


class _ClosingFailsHandle:
    """Wraps a real file handle, proxying everything but raising on close()."""

    def __init__(self, inner: Any) -> None:
        self._inner = inner

    def __iter__(self) -> Any:
        return iter(self._inner)

    def close(self) -> None:
        raise OSError("stale file handle")

    def __getattr__(self, name: str) -> Any:
        return getattr(self._inner, name)


def test_a_csv_read_failure_mid_file_is_a_value_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    # next(reader) pulls more bytes from the underlying file on every call, so
    # a removable/network filesystem failing mid-read raises a raw OSError —
    # not a csv.Error — which the field-limit-only translation above misses.
    path = _write_csv(tmp_path / "d.csv", [["a"], ["1"]])

    class FailingReader:
        line_num = 1

        def __iter__(self) -> FailingReader:
            return self

        def __next__(self) -> list[str]:
            raise OSError("input/output error")

    monkeypatch.setattr(csv, "reader", lambda *a, **k: FailingReader())
    with pytest.raises(ValueError, match="cannot read"):
        read_tabular_data(path)


def test_a_csv_close_failure_is_a_value_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = _write_csv(tmp_path / "d.csv", [["a"], ["1"]])
    real_open = Path.open

    def patched_open(self: Path, *args: object, **kwargs: object) -> Any:
        handle = real_open(self, *args, **kwargs)
        return _ClosingFailsHandle(handle) if self == path else handle

    monkeypatch.setattr(Path, "open", patched_open)
    with pytest.raises(ValueError, match="cannot read"):
        read_tabular_data(path)


def test_a_csv_close_failure_does_not_mask_a_parse_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    # A close() failure while a parse error is already propagating must not
    # replace that error with the close failure (the finally-block-exception
    # -replaces-try-exception trap).
    path = _write_csv(tmp_path / "d.csv", [["blob"], ["x" * (MAX_TABULAR_RESPONSE_BYTES + 10)]])
    real_open = Path.open

    def patched_open(self: Path, *args: object, **kwargs: object) -> Any:
        handle = real_open(self, *args, **kwargs)
        return _ClosingFailsHandle(handle) if self == path else handle

    monkeypatch.setattr(Path, "open", patched_open)
    with pytest.raises(ValueError, match="A single CSV field must stay under"):
        read_tabular_data(path)


# --- reading: malformed workbooks ----------------------------------------------------


def test_malformed_xlsx_raises_a_value_error(tmp_path: Path) -> None:
    # load_workbook raises zipfile.BadZipFile for a non-zip file — not a
    # ValueError — which would otherwise bypass the tools' error translation.
    target = tmp_path / "bad.xlsx"
    target.write_text("this is not a zip file", encoding="utf-8")
    with pytest.raises(ValueError, match="cannot read"):
        read_tabular_data(target)


def test_xlsx_with_corrupt_worksheet_xml_raises_a_value_error(tmp_path: Path) -> None:
    # read_only=True defers worksheet XML parsing until rows are iterated, so
    # load_workbook succeeds on a zip-valid file with corrupt worksheet XML;
    # the resulting ElementTree.ParseError surfaces only from iter_rows and
    # must be translated there too, not just at load_workbook.
    good = tmp_path / "good.xlsx"
    _write_xlsx(good, [["h1"], ["v1"]])
    target = tmp_path / "bad.xlsx"
    with zipfile.ZipFile(good, "r") as src, zipfile.ZipFile(target, "w") as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data = data[: len(data) // 2]
            dst.writestr(item, data)
    with pytest.raises(ValueError, match="cannot read"):
        read_tabular_data(target)


def test_xlsx_with_no_worksheets_raises_a_value_error(tmp_path: Path) -> None:
    # A zip-valid workbook with an empty <sheets> list makes load_workbook
    # succeed, but Workbook.active returns None (an IndexError swallowed
    # internally) rather than raising; dereferencing .title on that None
    # would otherwise leak an untranslated AttributeError.
    good = tmp_path / "good.xlsx"
    _write_xlsx(good, [["h1"], ["v1"]])
    target = tmp_path / "no_sheets.xlsx"
    with zipfile.ZipFile(good, "r") as src, zipfile.ZipFile(target, "w") as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "xl/workbook.xml":
                data = re.sub(rb"<sheets>.*?</sheets>", b"<sheets></sheets>", data)
            dst.writestr(item, data)
    with pytest.raises(ValueError, match="no worksheets"):
        read_tabular_data(target)


@pytest.mark.skipif(sys.platform == "win32", reason="POSIX permission bits")
@pytest.mark.skipif(
    hasattr(os, "geteuid") and os.geteuid() == 0, reason="root ignores permission bits"
)
def test_read_permission_denied_csv_is_a_value_error(tmp_path: Path) -> None:
    target = _write_csv(tmp_path / "in.csv", [["a"], ["1"]])
    target.chmod(0o000)
    try:
        with pytest.raises(ValueError, match="cannot read"):
            read_tabular_data(target)
    finally:
        target.chmod(0o644)
